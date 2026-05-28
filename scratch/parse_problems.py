import openpyxl
import os
import re

excel_path = r"c:\Users\DWIKY SUMARLIN\Documents\PORTOFOLIO\dji\DATA BASE SYSTEM (1).xlsx"

if not os.path.exists(excel_path):
    print(f"Error: {excel_path} not found")
    exit(1)

print("Loading workbook...")
wb = openpyxl.load_workbook(excel_path, read_only=True)
ws = wb["DATA BASE"]

# Extract headers
headers = [cell.value for cell in ws[1]]
print("Headers:", headers)

# Find column indices
operator_idx = -1
design_idx = -1
group_idx = -1
final_insp_idx = -1

problem_cols = {} # category_name -> column_index

for idx, h in enumerate(headers):
    if not h:
        continue
    h_clean = h.strip()
    if h_clean == "Nama Operator":
        operator_idx = idx
    elif h_clean == "Design":
        design_idx = idx
    elif h_clean == "Grup":
        group_idx = idx
    elif h_clean == "FInal Inspeksi":
        final_insp_idx = idx
    elif "MASALAH" in h_clean:
        # Category name: e.g. ELECTRIC, MEKANIK
        cat_name = h_clean.replace("MASALAH", "").strip()
        # Ensure we only track problem code columns, not description columns
        if not h_clean.startswith("Keterangan"):
            problem_cols[cat_name] = idx

print(f"Operator Col Index: {operator_idx}")
print(f"Design Col Index: {design_idx}")
print(f"Group Col Index: {group_idx}")
print(f"Final Inspection Col Index: {final_insp_idx}")
print(f"Problem Columns: {list(problem_cols.keys())}")

operators = set()
designs = set()
groups = set()
final_inspections = set()
problems_by_cat = {cat: set() for cat in problem_cols.keys()}

# Read all rows starting from row 2
row_num = 1
for row in ws.iter_rows(min_row=2, values_only=True):
    row_num += 1
    
    # Operators
    if operator_idx != -1 and row[operator_idx]:
        op = str(row[operator_idx]).strip()
        if op:
            operators.add(op)
            
    # Designs
    if design_idx != -1 and row[design_idx]:
        ds = str(row[design_idx]).strip()
        if ds:
            designs.add(ds)
            
    # Groups
    if group_idx != -1 and row[group_idx]:
        gr = str(row[group_idx]).strip()
        if gr:
            groups.add(gr)
            
    # Final Inspections
    if final_insp_idx != -1 and row[final_insp_idx]:
        fi = str(row[final_insp_idx]).strip()
        if fi:
            final_inspections.add(fi)
            
    # Problems
    for cat, col_idx in problem_cols.items():
        val = row[col_idx]
        if val:
            # Sometime comma separated, split them
            entries = re.split(r'[,;\n]+', str(val))
            for entry in entries:
                entry = entry.strip()
                if entry:
                    problems_by_cat[cat].add(entry)

print(f"\nUnique Operators ({len(operators)}):", sorted(list(operators)))
print(f"\nUnique Designs ({len(designs)}):", sorted(list(designs)))
print(f"\nUnique Groups ({len(groups)}):", sorted(list(groups)))
print(f"\nUnique Final Inspections ({len(final_inspections)}):", sorted(list(final_inspections)))

print("\nUnique Problems By Category:")
for cat, probs in problems_by_cat.items():
    print(f"  {cat} ({len(probs)} problems):")
    for p in sorted(list(probs)):
        print(f"    - {repr(p)}")
