import openpyxl
import os
import re
import json
import urllib.request
import urllib.error

# Paths
excel_path = r"c:\Users\DWIKY SUMARLIN\Documents\PORTOFOLIO\dji\DATA BASE SYSTEM (1).xlsx"
env_path = r"c:\Users\DWIKY SUMARLIN\Documents\PORTOFOLIO\dji\.env.local"
sql_output_path = r"c:\Users\DWIKY SUMARLIN\Documents\PORTOFOLIO\dji\scratch\seed.sql"

if not os.path.exists(excel_path):
    print(f"Error: Excel file not found at {excel_path}")
    exit(1)

# 1. Mappings Inisial Selaras dengan UI (Fallback)
operator_map = {
    "Rani": 1, "Rini": 2, "Neneng": 3, "Royana": 4, "Ridwan": 5, "Rina": 6, "Riki": 7, "Parid": 8, "Irfan": 9,
    "Sigit": 10, "Irma": 11, "Hardi": 12, "Gilang": 13, "Komara": 14, "Novi": 15, "Jaya": 16, "Ahmad": 17,
    "Rohmat": 18, "Devi": 19, "Anwar": 20, "Sandi": 21, "Yanti": 22, "Iki": 23
}

design_map = {
    "TCD 5826 XA": 1, "DL 5675 CO": 2, "DL 5167 CO": 3, "DL 5169 CO": 4, "DL 6460 CR": 5, "DL 5162 CO": 6, "DL 5168 CO": 7
}

group_map = {
    "A": 1, "B": 2, "C": 3
}

category_map = {
    "ELECTRIC": 1, "MEKANIK": 2, "ELEMENT RAJUTAN": 3, "BAHAN BAKU": 4, "MAINTENANCE/PERAWATAN": 5, "GANTI DESIGN": 6, "GANTI BENANG": 7, "MESIN STOP": 8
}

final_inspection_map = {
    "GRADE A": 1, "GRADE B": 2, "BS": 3
}

problem_fallback_list = [
    # ELECTRIC
    {"id": 1, "code": "A.1", "desc": "Mati Listrik", "cat": "ELECTRIC"},
    {"id": 2, "code": "A.3", "desc": "Error Servo Drive", "cat": "ELECTRIC"},
    {"id": 3, "code": "A.5", "desc": "Error Shogging", "cat": "ELECTRIC"},
    {"id": 4, "code": "A.6", "desc": "Error EBA", "cat": "ELECTRIC"},
    {"id": 5, "code": "A.7", "desc": "Error Jacquard", "cat": "ELECTRIC"},
    # MEKANIK
    {"id": 6, "code": "B.5", "desc": "Perbaikan tensioner", "cat": "MEKANIK"},
    # ELEMENT RAJUTAN
    {"id": 7, "code": "C.1", "desc": "Perbaikan jarum pattern patah/bengkok", "cat": "ELEMENT RAJUTAN"},
    {"id": 8, "code": "C.2", "desc": "Perbaikan Jacquard", "cat": "ELEMENT RAJUTAN"},
    {"id": 9, "code": "C.5", "desc": "Perbaikan Keluar Jarum", "cat": "ELEMENT RAJUTAN"},
    {"id": 10, "code": "C.7", "desc": "Perbaikan bolong corak", "cat": "ELEMENT RAJUTAN"},
    {"id": 11, "code": "C.9", "desc": "Perbaikan Ngegaris/Stopline", "cat": "ELEMENT RAJUTAN"},
    # BAHAN BAKU
    {"id": 12, "code": "D.5", "desc": "Perbaikan benang narik/kendor", "cat": "BAHAN BAKU"},
    {"id": 13, "code": "D.6", "desc": "Perbaikan benang nyilang", "cat": "BAHAN BAKU"},
    {"id": 14, "code": "D.7", "desc": "Perbaikan benang pinggiran", "cat": "BAHAN BAKU"},
    {"id": 15, "code": "D.8", "desc": "Perbaikan benang kusut", "cat": "BAHAN BAKU"},
    {"id": 16, "code": "D.9", "desc": "Perbaikan L1/2/3 putus", "cat": "BAHAN BAKU"},
    {"id": 17, "code": "D.10", "desc": "Beset L1/L2/L3", "cat": "BAHAN BAKU"},
    {"id": 18, "code": "D.13", "desc": "Benang timbul putus", "cat": "BAHAN BAKU"},
    # GANTI DESIGN
    {"id": 19, "code": "F.2", "desc": "Perbaikan corak/revisi", "cat": "GANTI DESIGN"},
]

problem_map = {} # (code, desc, cat) -> id
next_problem_id = 20

# Populate initial problems
for p in problem_fallback_list:
    key = (p["code"], p["desc"], p["cat"])
    problem_map[key] = p["id"]

# Load excel and extract any additions
print("Membuka workbook Excel...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["DATA BASE"]

headers = [cell.value for cell in ws[1]]

# Indeks Kolom
operator_idx = -1
design_idx = -1
group_idx = -1
final_insp_idx = -1
problem_cols = {} # category_name -> column_index

for idx, h in enumerate(headers):
    if not h:
        continue
    h_clean = h.strip()
    if h_clean == "Nama Operator":
        operator_idx = idx
    elif h_clean == "Design":
        design_idx = idx
    elif h_clean == "Grup":
        group_idx = idx
    elif h_clean == "FInal Inspeksi":
        final_insp_idx = idx
    elif "MASALAH" in h_clean and not h_clean.startswith("Keterangan"):
        cat_name = h_clean.replace("MASALAH", "").strip()
        problem_cols[cat_name] = idx

# Scan all rows to identify new operators, designs, groups, problems
print("Memindai data dari Excel...")
for row in ws.iter_rows(min_row=2, values_only=True):
    # Operator
    if operator_idx != -1 and row[operator_idx]:
        op = str(row[operator_idx]).strip()
        op_proper = op.title() # Convert case-insensitively
        # Match case-insensitively
        matched = False
        for k in operator_map.keys():
            if k.lower() == op.lower():
                matched = True
                break
        if not matched and op:
            new_id = max(operator_map.values()) + 1 if operator_map else 1
            operator_map[op_proper] = new_id

    # Design
    if design_idx != -1 and row[design_idx]:
        ds = str(row[design_idx]).strip()
        if ds and ds not in design_map:
            new_id = max(design_map.values()) + 1 if design_map else 1
            design_map[ds] = new_id

    # Group
    if group_idx != -1 and row[group_idx]:
        gr = str(row[group_idx]).strip()
        if gr and gr not in group_map:
            new_id = max(group_map.values()) + 1 if group_map else 1
            group_map[gr] = new_id

    # Final Inspection
    if final_insp_idx != -1 and row[final_insp_idx]:
        fi = str(row[final_insp_idx]).strip()
        if fi and fi not in final_inspection_map:
            new_id = max(final_inspection_map.values()) + 1 if final_inspection_map else 1
            final_inspection_map[fi] = new_id

    # Problems
    for cat, col_idx in problem_cols.items():
        val = row[col_idx]
        if val:
            entries = re.split(r'[,;\n]+', str(val))
            for entry in entries:
                entry = entry.strip()
                if not entry:
                    continue
                
                # Parse code and desc
                code = None
                desc = entry
                
                # Coba parsing code like C.5, D.9, A.1
                # Format 1: Code followed by tab or spaces
                parts = re.split(r'[\t]+', entry, maxsplit=1)
                if len(parts) == 2:
                    code, desc = parts[0].strip(), parts[1].strip()
                else:
                    # Format 2: Regex check for code like C.5 or D.10
                    m = re.match(r'^([A-Z]\.\d+)(?:\s+(.*))?$', entry)
                    if m:
                        code = m.group(1).strip()
                        desc = m.group(2).strip() if m.group(2) else ""
                    else:
                        code = None
                        desc = entry.strip()
                
                # Cek apakah sudah terpetakan
                found = False
                for (m_code, m_desc, m_cat), m_id in problem_map.items():
                    if m_cat == cat and m_code == code and m_desc.lower() == desc.lower():
                        found = True
                        break
                
                if not found:
                    problem_map[(code, desc, cat)] = next_problem_id
                    next_problem_id += 1

print(f"Total Operator Terpetakan: {len(operator_map)}")
print(f"Total Design Terpetakan: {len(design_map)}")
print(f"Total Grup Terpetakan: {len(group_map)}")
print(f"Total Final Inspeksi Terpetakan: {len(final_inspection_map)}")
print(f"Total Masalah Terpetakan: {len(problem_map)}")

# 2. Hasilkan SQL Seeder Premium
print("Membuat file SQL seeder...")
sql_lines = [
    "-- DJI Supabase Database Seeder Script",
    "-- Dihasilkan secara otomatis oleh Antigravity",
    "-- Catatan: Tabel lookup akan disemai terlebih dahulu dengan ID tetap agar sinkron dengan form Kios",
    "\nBEGIN;\n",
    "-- Nonaktifkan sementara pemicu jika ada",
    "SET session_replication_role = 'replica';\n"
]

# A. Seed Problem Categories
sql_lines.append("-- 1. SEED PROBLEM CATEGORIES")
for cat, cid in sorted(category_map.items(), key=lambda x: x[1]):
    sql_lines.append(f"INSERT INTO public.problem_categories (id, nama_kategori) VALUES ({cid}, '{cat}') ON CONFLICT (id) DO UPDATE SET nama_kategori = EXCLUDED.nama_kategori;")

# B. Seed Final Inspections
sql_lines.append("\n-- 2. SEED FINAL INSPECTIONS")
for fi, fid in sorted(final_inspection_map.items(), key=lambda x: x[1]):
    sql_lines.append(f"INSERT INTO public.final_inspections (id, status_final) VALUES ({fid}, '{fi}') ON CONFLICT (id) DO UPDATE SET status_final = EXCLUDED.status_final;")

# C. Seed Groups
sql_lines.append("\n-- 3. SEED GROUPS")
for gr, gid in sorted(group_map.items(), key=lambda x: x[1]):
    sql_lines.append(f"INSERT INTO public.groups (id, nama_grup) VALUES ({gid}, '{gr}') ON CONFLICT (id) DO UPDATE SET nama_grup = EXCLUDED.nama_grup;")

# D. Seed Designs
sql_lines.append("\n-- 4. SEED DESIGNS")
for ds, did in sorted(design_map.items(), key=lambda x: x[1]):
    # Escape single quote
    ds_escaped = ds.replace("'", "''")
    sql_lines.append(f"INSERT INTO public.designs (id, nama_design) VALUES ({did}, '{ds_escaped}') ON CONFLICT (id) DO UPDATE SET nama_design = EXCLUDED.nama_design;")

# E. Seed Operators
sql_lines.append("\n-- 5. SEED OPERATORS")
for op, oid in sorted(operator_map.items(), key=lambda x: x[1]):
    op_escaped = op.replace("'", "''")
    sql_lines.append(f"INSERT INTO public.operators (id, nama_operator) VALUES ({oid}, '{op_escaped}') ON CONFLICT (id) DO UPDATE SET nama_operator = EXCLUDED.nama_operator;")

# F. Seed Problems
sql_lines.append("\n-- 6. SEED PROBLEMS")
for (code, desc, cat), pid in sorted(problem_map.items(), key=lambda x: x[1]):
    cid = category_map[cat]
    code_val = f"'{code}'" if code else "NULL"
    desc_escaped = desc.replace("'", "''")
    sql_lines.append(f"INSERT INTO public.problems (id, kode_masalah, deskripsi_masalah, category_id) VALUES ({pid}, {code_val}, '{desc_escaped}', {cid}) ON CONFLICT (id) DO UPDATE SET kode_masalah = EXCLUDED.kode_masalah, deskripsi_masalah = EXCLUDED.deskripsi_masalah, category_id = EXCLUDED.category_id;")

# G. Reset Serial Sequences
sql_lines.append("\n-- 7. RESET SERIAL SEQUENCES")
sql_lines.extend([
    "SELECT setval(pg_get_serial_sequence('public.operators', 'id'), coalesce(max(id), 1)) FROM public.operators;",
    "SELECT setval(pg_get_serial_sequence('public.designs', 'id'), coalesce(max(id), 1)) FROM public.designs;",
    "SELECT setval(pg_get_serial_sequence('public.groups', 'id'), coalesce(max(id), 1)) FROM public.groups;",
    "SELECT setval(pg_get_serial_sequence('public.problem_categories', 'id'), coalesce(max(id), 1)) FROM public.problem_categories;",
    "SELECT setval(pg_get_serial_sequence('public.problems', 'id'), coalesce(max(id), 1)) FROM public.problems;",
    "SELECT setval(pg_get_serial_sequence('public.final_inspections', 'id'), coalesce(max(id), 1)) FROM public.final_inspections;"
])

sql_lines.append("\n-- Aktifkan kembali pemicu")
sql_lines.append("SET session_replication_role = 'origin';")
sql_lines.append("\nCOMMIT;")

# Tulis ke file seed.sql
os.makedirs(os.path.dirname(sql_output_path), exist_ok=True)
with open(sql_output_path, "w", encoding="utf-8") as f:
    f.write("\n".join(sql_lines))
print(f"Sukses membuat file SQL seeder di: {sql_output_path}")

# 3. Coba Hubungkan Langsung ke Supabase via HTTP REST API jika ada kredensial di .env.local
supabase_url = None
supabase_service_role = None

if os.path.exists(env_path):
    print("Membaca konfigurasi .env.local...")
    with open(env_path, "r", encoding="utf-8") as f:
        content = f.read()
        url_match = re.search(r"NEXT_PUBLIC_SUPABASE_URL\s*=\s*(https://[^\s#]+)", content)
        if url_match:
            supabase_url = url_match.group(1).strip()
        
        role_match = re.search(r"SUPABASE_SERVICE_ROLE_KEY\s*=\s*([^\s#]+)", content)
        if role_match:
            supabase_service_role = role_match.group(1).strip()

if (supabase_url and supabase_service_role and 
    "your_supabase_anon_key_here" not in supabase_service_role and 
    "your_supabase_service_role_key_here" not in supabase_service_role):
    
    print(f"Ditemukan kredensial Supabase! URL: {supabase_url}")
    print("Memulai unggah data lookups langsung ke Supabase...")
    
    def supabase_upsert(table_name, records):
        url = f"{supabase_url}/rest/v1/{table_name}"
        headers = {
            "apikey": supabase_service_role,
            "Authorization": f"Bearer {supabase_service_role}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates"
        }
        data = json.dumps(records).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req) as resp:
                print(f"  [+] Berhasil mengunggah {len(records)} record ke tabel {table_name}")
                return True
        except urllib.error.HTTPError as e:
            print(f"  [-] Gagal mengunggah ke {table_name}: HTTP {e.code} - {e.read().decode('utf-8')}")
            return False
        except Exception as ex:
            print(f"  [-] Gagal mengunggah ke {table_name}: {ex}")
            return False

    # A. Upsert Problem Categories
    cat_records = [{"id": cid, "nama_kategori": cat} for cat, cid in category_map.items()]
    supabase_upsert("problem_categories", cat_records)

    # B. Upsert Final Inspections
    fi_records = [{"id": fid, "status_final": fi} for fi, fid in final_inspection_map.items()]
    supabase_upsert("final_inspections", fi_records)

    # C. Upsert Groups
    gr_records = [{"id": gid, "nama_grup": gr} for gr, gid in group_map.items()]
    supabase_upsert("groups", gr_records)

    # D. Upsert Designs
    ds_records = [{"id": did, "nama_design": ds} for ds, did in design_map.items()]
    supabase_upsert("designs", ds_records)

    # E. Upsert Operators
    op_records = [{"id": oid, "nama_operator": op} for op, oid in operator_map.items()]
    supabase_upsert("operators", op_records)

    # F. Upsert Problems
    prob_records = []
    for (code, desc, cat), pid in problem_map.items():
        prob_records.append({
            "id": pid,
            "kode_masalah": code,
            "deskripsi_masalah": desc,
            "category_id": category_map[cat]
        })
    supabase_upsert("problems", prob_records)

    print("\n[OK] Proses unggah database selesai!")
else:
    print("\n[!] Kredensial Supabase di .env.local belum diisi atau masih menggunakan placeholder.")
    print("    Silakan jalankan script ini kembali setelah Anda mengonfigurasi Anon & Service Role Key Anda.")
    print("    ATAU salin isi file SQL yang dihasilkan di bawah ke Supabase SQL Editor:")
    print(f"    -> [seed.sql]({sql_output_path})")

print("\n--- Selesai ---")
